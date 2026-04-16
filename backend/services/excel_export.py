"""Multi-sheet Excel workbook generation using openpyxl."""

import io
import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

from services.query_engine import query_receipts

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
CURRENCY_FMT = '#,##0.00'


def _style_header(ws, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


def _apply_alt_rows(ws, start_row: int, end_row: int, col_count: int):
    for row in range(start_row, end_row + 1):
        if row % 2 == 0:
            for col in range(1, col_count + 1):
                ws.cell(row=row, column=col).fill = ALT_ROW_FILL


def _format_currency_cols(ws, cols: list[int], start_row: int, end_row: int):
    for row in range(start_row, end_row + 1):
        for col in cols:
            cell = ws.cell(row=row, column=col)
            cell.number_format = CURRENCY_FMT
            cell.alignment = Alignment(horizontal="right")


def generate_excel(conn: sqlite3.Connection, **filters) -> io.BytesIO:
    """Generate a multi-sheet Excel workbook. Returns bytes buffer."""
    _, receipts = query_receipts(conn, **filters)

    wb = Workbook()

    # --- Sheet 1: Raw_Data ---
    ws = wb.active
    ws.title = "Raw_Data"
    headers = ["ID", "Batch", "Vendor", "Normalized Vendor", "Date", "Invoice No",
               "Amount", "Tax", "Total", "Category", "Hostel", "Account No", "IFSC", "Remarks"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for r in receipts:
        ws.append([
            r["id"], r["batch_id"], r["vendor"], r["normalized_vendor"],
            r["date"], r["invoice_no"], r["amount"], r["tax"], r["total"],
            r["category"], r["hostel_no"], r["account_no"], r["ifsc"], r["remarks"],
        ])

    _apply_alt_rows(ws, 2, len(receipts) + 1, len(headers))
    _format_currency_cols(ws, [7, 8, 9], 2, len(receipts) + 1)
    _auto_width(ws)

    # --- Sheet 2: Vendor_Summary ---
    ws2 = wb.create_sheet("Vendor_Summary")
    vendor_agg = {}
    for r in receipts:
        v = r["normalized_vendor"] or "UNKNOWN"
        if v not in vendor_agg:
            vendor_agg[v] = {"amount": 0, "tax": 0, "total": 0, "count": 0}
        vendor_agg[v]["amount"] += r["amount"] or 0
        vendor_agg[v]["tax"] += r["tax"] or 0
        vendor_agg[v]["total"] += r["total"] or 0
        vendor_agg[v]["count"] += 1

    ws2.append(["Vendor", "Total Amount", "Total Tax", "Grand Total", "Count"])
    _style_header(ws2, 5)
    row_num = 2
    for v in sorted(vendor_agg.keys()):
        d = vendor_agg[v]
        ws2.append([v, round(d["amount"], 2), round(d["tax"], 2), round(d["total"], 2), d["count"]])
        row_num += 1
    _apply_alt_rows(ws2, 2, row_num - 1, 5)
    _format_currency_cols(ws2, [2, 3, 4], 2, row_num - 1)
    _auto_width(ws2)

    # --- Sheet 3: Category_Summary ---
    ws3 = wb.create_sheet("Category_Summary")
    cat_agg = {}
    for r in receipts:
        c = r["category"] or "Uncategorized"
        if c not in cat_agg:
            cat_agg[c] = {"total": 0, "count": 0}
        cat_agg[c]["total"] += r["total"] or 0
        cat_agg[c]["count"] += 1

    ws3.append(["Category", "Total Amount", "Count"])
    _style_header(ws3, 3)
    row_num = 2
    for c in sorted(cat_agg.keys()):
        ws3.append([c, round(cat_agg[c]["total"], 2), cat_agg[c]["count"]])
        row_num += 1
    _apply_alt_rows(ws3, 2, row_num - 1, 3)
    _format_currency_cols(ws3, [2], 2, row_num - 1)
    _auto_width(ws3)

    # --- Hostel sheets ---
    hostel_data = {}
    for r in receipts:
        h = r["hostel_no"]
        if h is None:
            continue
        if h not in hostel_data:
            hostel_data[h] = {}
        v = r["normalized_vendor"] or "UNKNOWN"
        if v not in hostel_data[h]:
            hostel_data[h][v] = {"account_no": r["account_no"], "ifsc": r["ifsc"], "total": 0, "remarks": r["remarks"] or ""}
        hostel_data[h][v]["total"] += r["total"] or 0
        if r["account_no"]:
            hostel_data[h][v]["account_no"] = r["account_no"]
        if r["ifsc"]:
            hostel_data[h][v]["ifsc"] = r["ifsc"]

    for h in sorted(hostel_data.keys()):
        ws_h = wb.create_sheet(f"Hostel_{h}")
        ws_h.append(["Sr", "Vendor Name", "Account No", "IFSC", "Amount", "Remarks"])
        _style_header(ws_h, 6)
        row_num = 2
        grand = 0
        for i, v in enumerate(sorted(hostel_data[h].keys()), 1):
            d = hostel_data[h][v]
            ws_h.append([i, v, d["account_no"], d["ifsc"], round(d["total"], 2), d["remarks"]])
            grand += d["total"]
            row_num += 1
        # Total row
        ws_h.append(["", "", "", "TOTAL", round(grand, 2), ""])
        ws_h.cell(row=row_num, column=4).font = Font(bold=True)
        ws_h.cell(row=row_num, column=5).font = Font(bold=True)
        _apply_alt_rows(ws_h, 2, row_num - 1, 6)
        _format_currency_cols(ws_h, [5], 2, row_num, )
        _auto_width(ws_h)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
